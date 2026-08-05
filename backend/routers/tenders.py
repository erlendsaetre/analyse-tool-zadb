"""
Tenders Router
==============
Handles all CRUD for Tenders, TenderRates, and file imports.

Import formats supported:
  - kn_row_based: KN TE Connect / standard KN tender export (Row_based sheet, headers row 12, data from row 14)

Future import formats can be added by extending _parse_kn_row_based() or adding new parser functions
and wiring them in the import endpoint via the format_type parameter.
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from typing import List, Optional
from database import get_db
import models, schemas
import openpyxl
import io
from datetime import datetime

router = APIRouter()


# ══════════════════════════════════════════════════════════════
#  Tender CRUD
# ══════════════════════════════════════════════════════════════

@router.get("/", response_model=List[schemas.TenderListItem])
def list_tenders(db: Session = Depends(get_db)):
    tenders = db.query(models.Tender).order_by(models.Tender.updated_at.desc()).all()
    result = []
    for t in tenders:
        result.append(schemas.TenderListItem(
            id=t.id,
            name=t.name,
            description=t.description,
            customer=t.customer,
            status=t.status,
            tender_url=t.tender_url,
            valid_until=t.valid_until,
            created_at=t.created_at,
            updated_at=t.updated_at,
            rate_count=len(t.rates)
        ))
    return result


@router.post("/", response_model=schemas.TenderResponse)
def create_tender(tender: schemas.TenderCreate, db: Session = Depends(get_db)):
    db_tender = models.Tender(
        name=tender.name,
        description=tender.description,
        customer=tender.customer,
    )
    db.add(db_tender)
    db.commit()
    db.refresh(db_tender)
    return db_tender


@router.get("/{tender_id}", response_model=schemas.TenderResponse)
def get_tender(tender_id: int, db: Session = Depends(get_db)):
    tender = db.query(models.Tender).filter(models.Tender.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")
    return tender


@router.put("/{tender_id}", response_model=schemas.TenderResponse)
def update_tender(tender_id: int, update: schemas.TenderUpdate, db: Session = Depends(get_db)):
    tender = db.query(models.Tender).filter(models.Tender.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")
    update_data = update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(tender, key, value)
    db.commit()
    db.refresh(tender)
    return tender


@router.delete("/{tender_id}")
def delete_tender(tender_id: int, db: Session = Depends(get_db)):
    tender = db.query(models.Tender).filter(models.Tender.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")
    db.delete(tender)
    db.commit()
    return {"message": f"Tender '{tender.name}' deleted"}


# ══════════════════════════════════════════════════════════════
#  TenderRate CRUD
# ══════════════════════════════════════════════════════════════

@router.post("/{tender_id}/rates", response_model=schemas.TenderRateResponse)
def add_rate_to_tender(tender_id: int, rate: schemas.TenderRateCreate, db: Session = Depends(get_db)):
    tender = db.query(models.Tender).filter(models.Tender.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")
    db_rate = models.TenderRate(tender_id=tender_id, **rate.model_dump())
    db.add(db_rate)
    db.commit()
    db.refresh(db_rate)
    return db_rate


@router.put("/{tender_id}/rates/{rate_id}", response_model=schemas.TenderRateResponse)
def update_tender_rate(tender_id: int, rate_id: int, update: schemas.TenderRateUpdate, db: Session = Depends(get_db)):
    rate = db.query(models.TenderRate).filter(
        models.TenderRate.id == rate_id,
        models.TenderRate.tender_id == tender_id
    ).first()
    if not rate:
        raise HTTPException(status_code=404, detail="Rate not found in this tender")
    update_data = update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(rate, key, value)
    db.commit()
    db.refresh(rate)
    return rate


@router.delete("/{tender_id}/rates/{rate_id}")
def delete_tender_rate(tender_id: int, rate_id: int, db: Session = Depends(get_db)):
    rate = db.query(models.TenderRate).filter(
        models.TenderRate.id == rate_id,
        models.TenderRate.tender_id == tender_id
    ).first()
    if not rate:
        raise HTTPException(status_code=404, detail="Rate not found in this tender")
    db.delete(rate)
    db.commit()
    return {"message": "Rate removed from tender"}


# ══════════════════════════════════════════════════════════════
#  File Import — into existing tender
# ══════════════════════════════════════════════════════════════

@router.post("/{tender_id}/import")
async def import_file_into_tender(
    tender_id: int,
    file: UploadFile = File(...),
    format_type: str = Form("kn_row_based"),
    import_notes: str = Form(None),
    db: Session = Depends(get_db)
):
    """
    Import an Excel file into an EXISTING tender.
    
    Supported format_type values:
      - kn_row_based (default): KN TE Connect / standard KN tender export
    
    Returns a summary of what was imported.
    """
    tender = db.query(models.Tender).filter(models.Tender.id == tender_id).first()
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")

    if not file.filename.endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xls, .xlsx) are supported.")

    try:
        contents = await file.read()

        if format_type == "kn_row_based":
            rates_data, parse_notes = _parse_kn_row_based(contents)
        else:
            raise HTTPException(status_code=400, detail=f"Unknown format_type: {format_type}")

        # Record this import
        import_record = models.TenderImport(
            tender_id=tender_id,
            filename=file.filename,
            format_type=format_type,
            imported_at=datetime.utcnow(),
            lane_count=len(rates_data),
            notes=import_notes or parse_notes
        )
        db.add(import_record)
        db.flush()

        # Insert all extracted rates
        for rd in rates_data:
            db_rate = models.TenderRate(
                tender_id=tender_id,
                tender_import_id=import_record.id,
                **rd
            )
            db.add(db_rate)

        db.commit()

        return {
            "message": f"Imported {len(rates_data)} lanes from '{file.filename}'",
            "tender_id": tender_id,
            "import_id": import_record.id,
            "lane_count": len(rates_data),
            "format": format_type,
            "parse_notes": parse_notes
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error parsing file: {str(e)}")


# ══════════════════════════════════════════════════════════════
#  Parser: KN Row_based format
# ══════════════════════════════════════════════════════════════

def _col_letter_to_index(col: str) -> int:
    """Convert Excel column letter(s) to 1-based column index."""
    result = 0
    for ch in col.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


# Column definitions for the KN "Row_based" sheet format.
# Key: our field name, Value: Excel column letter (headers on row 12, data from row 14)
# To adapt for a new export format, create a new COL_MAP and parser function.
KN_ROW_BASED_COLS = {
    'lane_id':       'H',   # KN Lane ID
    'origin_city':   'S',   # Origin City (for notes)
    'origin':        'X',   # KN Assigned Origin Airport
    'dest_city':     'AJ',  # Destination City (for notes)
    'destination':   'AO',  # KN Assigned Destination Airport
    'service_level': 'BM',  # KN Service Level
    'product':       'BN',  # KN Product
    'terms':         'BO',  # Terms of Delivery
    'currency':      'DE',  # Main Carriage Currency
    'cost_min':      'DM',  # Main Carriage MIN
    'cost_normal':   'DN',  # Main Carriage +0KG
    'cost_q45':      'DO',  # Main Carriage +45KG
    'cost_q100':     'DP',  # Main Carriage +100KG
    'cost_q300':     'DQ',  # Main Carriage +300KG
    'cost_q500':     'DR',  # Main Carriage +500KG
    'cost_q1000':    'DS',  # Main Carriage +1000KG
    'cost_q3000':    'DT',  # Main Carriage +3000KG
    'airline':       'EB',  # Carrier (IATA codes, may be comma-separated)
    'routing':       'EC',  # Routing e.g. OSL-DOH-BKK
    'via':           'ED',  # Transit Airport
}


def _parse_kn_row_based(contents: bytes):
    """
    Parse a KN TE Connect / Row_based format Excel file.
    Returns (list_of_rate_dicts, parse_notes_string).
    """
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)

    # Find the sheet
    sheet_name = None
    for name in wb.sheetnames:
        if 'row_based' in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    # Convert column letters to indices once
    col_idx = {field: _col_letter_to_index(col) for field, col in KN_ROW_BASED_COLS.items()}

    def cell(row, field):
        return ws.cell(row=row, column=col_idx[field]).value

    def sf(val):
        if val is None: return None
        try: return float(val)
        except: return None

    def ss(val):
        if val is None: return ""
        return str(val).strip()

    rates = []
    skipped = 0

    # Data rows start at row 14
    for row_idx in range(14, ws.max_row + 1):
        lane_id = cell(row_idx, 'lane_id')
        if not lane_id:
            skipped += 1
            continue

        airline = ss(cell(row_idx, 'airline'))
        origin = ss(cell(row_idx, 'origin'))
        destination = ss(cell(row_idx, 'destination'))
        routing = ss(cell(row_idx, 'routing'))
        via = ss(cell(row_idx, 'via'))
        product = ss(cell(row_idx, 'product'))
        service_level = ss(cell(row_idx, 'service_level'))
        terms = ss(cell(row_idx, 'terms'))
        currency = ss(cell(row_idx, 'currency')) or 'NOK'
        origin_city = ss(cell(row_idx, 'origin_city'))
        dest_city = ss(cell(row_idx, 'dest_city'))

        product_label = product
        if service_level and service_level != product:
            product_label = f"{product} ({service_level})"

        notes_parts = [f"Lane {lane_id}"]
        if origin_city: notes_parts.append(f"From: {origin_city}")
        if dest_city: notes_parts.append(f"To: {dest_city}")
        if routing: notes_parts.append(f"Route: {routing}")

        rates.append({
            'lane_id': str(lane_id),
            'airline': airline or 'Unknown',
            'product': product_label,
            'origin': origin,
            'destination': destination,
            'via': via or None,
            'routing': routing or None,
            'currency': currency,
            'terms': terms or None,
            'cost_min': sf(cell(row_idx, 'cost_min')),
            'cost_normal': sf(cell(row_idx, 'cost_normal')),
            'cost_q45': sf(cell(row_idx, 'cost_q45')),
            'cost_q100': sf(cell(row_idx, 'cost_q100')),
            'cost_q300': sf(cell(row_idx, 'cost_q300')),
            'cost_q500': sf(cell(row_idx, 'cost_q500')),
            'cost_q1000': sf(cell(row_idx, 'cost_q1000')),
            'cost_q3000': sf(cell(row_idx, 'cost_q3000')),
            'notes': ' | '.join(notes_parts),
        })

    parse_notes = f"Sheet: '{sheet_name}'. Imported {len(rates)} lanes, skipped {skipped} empty rows."
    return rates, parse_notes
