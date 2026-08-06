"""
Smart Excel Parser
==================
Auto-detects Excel file format and extracts ALL information intelligently.
Nothing is ever lost — unmapped columns go into a metadata dict.

Supported formats (auto-detected):
  1. KN Row_based: KN TE Connect / tender export with "Row_based" sheet
  2. KN ML-Rates: Standard KN quote file with "ML-Rates" sheet
  3. Generic: Falls back to header-matching on any sheet
"""

import openpyxl
import pandas as pd
import io
import re
import json
from datetime import datetime


# ══════════════════════════════════════════════════════════════
#  Known header patterns → our internal field names
#  Everything NOT mapped here goes into metadata JSON.
# ══════════════════════════════════════════════════════════════

# Maps known header text (lowercase, stripped) to our internal DB field names.
HEADER_MAP = {
    # ─── Airline / Carrier ───
    'airline': 'airline',
    'carrier': 'airline',
    'carrier (iata codes)': 'airline',
    'iata code': 'airline',
    
    # ─── GSA ───
    'gsa': 'gsa',
    
    # ─── Product / Service ───
    'product': 'product',
    'service': 'product',
    'service level': 'product',
    'kn product': 'product',
    'kn service level': 'product',
    
    # ─── Origin (airport code) ───
    'origin': 'origin',
    'origin airport': 'origin',
    'kn assigned origin airport': 'origin',
    'kn assigned origin airport [3 letter code]': 'origin',
    'client origin airport [3 letter code]': 'origin',
    'pol': 'origin',
    
    # ─── Destination (airport code) ───
    'destination': 'destination',
    'dest': 'destination',
    'destination airport': 'destination',
    'kn assigned destination airport': 'destination',
    'kn assigned destination airport [3 letter code]': 'destination',
    'client destination airport [3 letter code]': 'destination',
    'pod': 'destination',
    
    # ─── Origin details ───
    'origin city': 'origin_city',
    'origin city name': 'origin_city',
    'origin city [full name]': 'origin_city',
    'origin country': 'origin_country',
    'origin country code': 'origin_country',
    'origin country code [2 letter code]': 'origin_country',
    'origin country name [full name]': 'origin_country',
    'origin city zip code': 'origin_zip',
    'pickup zip': 'origin_zip',
    'pickup postal code': 'origin_zip',
    'pickup zip code': 'origin_zip',
    'collection zip': 'origin_zip',
    'kn assigned origin gateway': 'origin_gateway',
    'kn assigned origin gateway [3 letter code]': 'origin_gateway',
    
    # ─── Destination details ───
    'destination city': 'destination_city',
    'dest city': 'destination_city',
    'destination city name': 'destination_city',
    'destination city [full name]': 'destination_city',
    'destination country': 'destination_country',
    'destination country code': 'destination_country',
    'destination country code [2 letter code]': 'destination_country',
    'destination country name [full name]': 'destination_country',
    'destination city zip code': 'destination_zip',
    'delivery zip': 'destination_zip',
    'delivery postal code': 'destination_zip',
    'delivery zip code': 'destination_zip',
    'kn assigned destination gateway': 'destination_gateway',
    'kn assigned destination gateway [3 letter code]': 'destination_gateway',
    
    # ─── Via / Routing ───
    'via': 'via',
    'transit': 'via',
    'transit airport': 'via',
    'transit airport [3 letter code]': 'via',
    'routing': 'routing',
    'route': 'routing',
    
    # ─── Rate brackets ───
    'min': 'cost_min',
    'minimum': 'cost_min',
    'min rate': 'cost_min',
    'main carriage min': 'cost_min',
    'main carriage (min)': 'cost_min',
    
    'normal': 'cost_normal',
    '+0kg': 'cost_normal',
    'normal rate': 'cost_normal',
    'main carriage +0kg': 'cost_normal',
    'main carriage (+0kg)': 'cost_normal',
    
    'q45': 'cost_q45',
    '+45': 'cost_q45',
    '+45kg': 'cost_q45',
    'main carriage +45kg': 'cost_q45',
    'main carriage (+45kg)': 'cost_q45',
    
    'q100': 'cost_q100',
    '+100': 'cost_q100',
    '+100kg': 'cost_q100',
    'main carriage +100kg': 'cost_q100',
    'main carriage (+100kg)': 'cost_q100',
    
    'q300': 'cost_q300',
    '+300': 'cost_q300',
    '+300kg': 'cost_q300',
    'main carriage +300kg': 'cost_q300',
    'main carriage (+300kg)': 'cost_q300',
    
    'q500': 'cost_q500',
    '+500': 'cost_q500',
    '+500kg': 'cost_q500',
    'main carriage +500kg': 'cost_q500',
    'main carriage (+500kg)': 'cost_q500',
    
    'q1000': 'cost_q1000',
    '+1000': 'cost_q1000',
    '+1000kg': 'cost_q1000',
    'main carriage +1000kg': 'cost_q1000',
    'main carriage (+1000kg)': 'cost_q1000',
    
    'q3000': 'cost_q3000',
    '+3000': 'cost_q3000',
    '+3000kg': 'cost_q3000',
    'main carriage +3000kg': 'cost_q3000',
    'main carriage (+3000kg)': 'cost_q3000',
    
    # ─── Currency ───
    'currency': 'currency',
    'curr': 'currency',
    'ccy': 'currency',
    'main carriage currency': 'currency',
    
    # ─── Validity ───
    'valid': 'valid_from',
    'valid from': 'valid_from',
    'start date': 'valid_from',
    'effective date': 'valid_from',
    'effective from': 'valid_from',
    
    'expires': 'valid_until',
    'valid until': 'valid_until',
    'valid to': 'valid_until',
    'expiry': 'valid_until',
    'expiry date': 'valid_until',
    'end date': 'valid_until',
    
    # ─── Fuel / Surcharges ───
    'fuel': 'fuel',
    'fuel surcharge': 'fuel',
    'security surcharge': 'security_surcharge',
    'security': 'security_surcharge',
    'fixed': 'fixed',
    
    # ─── Volume / Weight ───
    'relation kg/m3': 'relation_kg_m3',
    'volume ratio': 'relation_kg_m3',
    
    # ─── Terms ───
    'terms': 'terms',
    'terms of delivery': 'terms',
    'incoterm': 'terms',
    'incoterms': 'terms',
    
    # ─── Direction ───
    'movement type': 'direction',
    'traffic direction': 'direction',
    
    # ─── Lane ID ───
    'kn lane id': 'lane_id',
    'lane id': 'lane_id',
    'lane': 'lane_id',
}

# Fields that are stored as dedicated DB columns (not in metadata)
DB_FIELDS = {
    'airline', 'product', 'origin', 'destination', 'via', 'routing',
    'currency', 'terms', 'lane_id',
    'origin_city', 'origin_country', 'origin_zip', 'origin_gateway',
    'destination_city', 'destination_country', 'destination_zip', 'destination_gateway',
    'cost_min', 'cost_normal', 'cost_q45', 'cost_q100', 'cost_q300',
    'cost_q500', 'cost_q1000', 'cost_q3000',
    'valid_from', 'valid_until',
}


def _safe_float(val):
    if val is None: return None
    if isinstance(val, (int, float)):
        return float(val) if val == val else None
    s = str(val).strip()
    if s.lower() in ('', '-', 'n/a', 'na', 'no', 'nan', 'o.r', 'o.r.', 'on request'):
        return None
    s = s.replace(',', '')
    try: return float(s)
    except (ValueError, TypeError): return None


def _safe_str(val):
    if val is None: return ""
    if isinstance(val, float):
        if val != val: return ""
        if val == int(val): return str(int(val))
    s = str(val).strip()
    if s.lower() in ('nan', 'none'): return ""
    return s


def _safe_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val
    s = str(val).strip()
    if s.lower() in ('', '-', 'n/a', 'na', 'nan'): return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d.%m.%Y', '%d-%m-%Y', '%m/%d/%Y'):
        try: return datetime.strptime(s, fmt)
        except ValueError: continue
    try: return pd.to_datetime(s, dayfirst=True).to_pydatetime()
    except Exception: return None


# ══════════════════════════════════════════════════════════════
#  Main entry point
# ══════════════════════════════════════════════════════════════

def smart_parse(contents: bytes):
    """
    Auto-detect Excel format and parse ALL data.
    Returns (rates: list[dict], metadata: dict)
    
    Each rate dict contains DB fields + 'metadata' key with JSON of all extra columns.
    """
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    
    # ── Step 1: Detect format ──
    sheet_names_lower = {name.lower(): name for name in wb.sheetnames}
    format_detected = 'generic'
    target_sheet = None
    
    for name_lower, name_actual in sheet_names_lower.items():
        if 'row_based' in name_lower:
            format_detected = 'kn_row_based'
            target_sheet = name_actual
            break
    
    if not target_sheet and 'ml-rates' in sheet_names_lower:
        format_detected = 'ml_rates'
        target_sheet = sheet_names_lower['ml-rates']
    
    if not target_sheet:
        for name in wb.sheetnames:
            if name.lower() != 'general':
                target_sheet = name
                break
        if not target_sheet:
            target_sheet = wb.sheetnames[0]
    
    ws = wb[target_sheet]
    
    # ── Step 2: Find header row ──
    header_row = _find_header_row(ws, format_detected)
    
    # ── Step 3: Map ALL columns ──
    column_mapping, unmapped_columns = _map_all_columns(ws, header_row)
    
    # ── Step 4: Detect direction ──
    direction = _detect_direction(ws, column_mapping, header_row)
    
    # ── Step 5: Extract data ──
    data_start = header_row + 1
    if format_detected == 'kn_row_based':
        data_start = header_row + 2
    
    rates = []
    skipped = 0
    
    for row_idx in range(data_start, ws.max_row + 1):
        # Check if entire row is empty
        has_any = False
        for col_idx in list(column_mapping.values()) + [v[1] for v in unmapped_columns]:
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None and str(val).strip() not in ('', 'nan'):
                has_any = True
                break
        
        if not has_any:
            skipped += 1
            continue
        
        # Extract mapped fields
        row_data = {}
        for our_field, col_idx in column_mapping.items():
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if our_field.startswith('cost_') or our_field == 'relation_kg_m3':
                row_data[our_field] = _safe_float(cell_val)
            elif our_field in ('valid_from', 'valid_until'):
                row_data[our_field] = _safe_date(cell_val)
            else:
                row_data[our_field] = _safe_str(cell_val)
        
        # Extract ALL unmapped columns into metadata dict
        extra = {}
        for header_text, col_idx in unmapped_columns:
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                val_str = _safe_str(cell_val)
                if val_str:  # Don't store empty strings
                    extra[header_text] = val_str
        
        # Must have at least lane_id, airline, or origin to be a valid row
        if not row_data.get('lane_id') and not row_data.get('airline') and not row_data.get('origin'):
            skipped += 1
            continue
        
        # Build the rate dict with DB fields
        rate = {
            'lane_id': row_data.get('lane_id') or None,
            'airline': row_data.get('airline') or None,
            'product': row_data.get('product') or None,
            'origin': row_data.get('origin') or None,
            'destination': row_data.get('destination') or None,
            'via': row_data.get('via') or None,
            'routing': row_data.get('routing') or None,
            'currency': row_data.get('currency') or None,
            'terms': row_data.get('terms') or None,
            'origin_city': row_data.get('origin_city') or None,
            'origin_country': row_data.get('origin_country') or None,
            'origin_zip': row_data.get('origin_zip') or None,
            'origin_gateway': row_data.get('origin_gateway') or None,
            'destination_city': row_data.get('destination_city') or None,
            'destination_country': row_data.get('destination_country') or None,
            'destination_zip': row_data.get('destination_zip') or None,
            'destination_gateway': row_data.get('destination_gateway') or None,
            'cost_min': row_data.get('cost_min'),
            'cost_normal': row_data.get('cost_normal'),
            'cost_q45': row_data.get('cost_q45'),
            'cost_q100': row_data.get('cost_q100'),
            'cost_q300': row_data.get('cost_q300'),
            'cost_q500': row_data.get('cost_q500'),
            'cost_q1000': row_data.get('cost_q1000'),
            'cost_q3000': row_data.get('cost_q3000'),
            'valid_from': row_data.get('valid_from'),
            'valid_until': row_data.get('valid_until'),
            'notes': None,
            'extra_data': json.dumps(extra, ensure_ascii=False) if extra else None,
        }
        
        # Also include direction and any other mapped-but-not-DB fields in extra_data
        mapped_extra = {}
        for field in ('direction', 'fuel', 'security_surcharge', 'fixed', 'relation_kg_m3', 'gsa'):
            val = row_data.get(field)
            if val:
                mapped_extra[field] = str(val) if not isinstance(val, str) else val
        if mapped_extra:
            existing = json.loads(rate['extra_data']) if rate['extra_data'] else {}
            existing.update(mapped_extra)
            rate['extra_data'] = json.dumps(existing, ensure_ascii=False)
        
        rates.append(rate)
    
    # Build metadata summary
    mapped_summary = {k: _col_index_to_letter(v) for k, v in column_mapping.items()}
    unmapped_summary = [f"{_col_index_to_letter(col_idx)}: {h}" for h, col_idx in unmapped_columns]
    
    metadata = {
        'format_detected': format_detected,
        'sheet_used': target_sheet,
        'header_row': header_row,
        'columns_mapped': mapped_summary,
        'columns_unmapped': unmapped_summary,
        'direction': direction,
        'total_rows': len(rates),
        'skipped_rows': skipped,
        'notes': (f"Format: {format_detected} | Sheet: '{target_sheet}' | "
                  f"Headers on row {header_row} | {len(rates)} rates extracted, "
                  f"{skipped} rows skipped | Direction: {direction} | "
                  f"Mapped {len(column_mapping)} columns, {len(unmapped_columns)} extra in metadata")
    }
    
    return rates, metadata


# ══════════════════════════════════════════════════════════════
#  Helper functions
# ══════════════════════════════════════════════════════════════

def _find_header_row(ws, format_detected: str) -> int:
    if format_detected == 'kn_row_based':
        return 12
    
    header_keywords = {'airline', 'carrier', 'origin', 'destination', 'product',
                       'min', 'normal', 'q100', 'currency', '+100', 'via'}
    best_row = 1
    best_score = 0
    for row_idx in range(1, min(21, ws.max_row + 1)):
        score = 0
        for col_idx in range(1, min(50, ws.max_column + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and str(val).strip().lower() in header_keywords:
                score += 1
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row


def _map_all_columns(ws, header_row: int):
    """
    Map ALL columns. Returns:
      - mapping: dict[str, int]  — our field name → column index (for DB fields)
      - unmapped: list[(str, int)]  — (original_header, col_index) for extra columns
    """
    mapping = {}
    unmapped = []
    
    for col_idx in range(1, ws.max_column + 1):
        header_val = ws.cell(row=header_row, column=col_idx).value
        if header_val is None:
            continue
        
        header_text = str(header_val).strip()
        header_lower = header_text.lower()
        if not header_lower:
            continue
        
        # Try exact match first
        if header_lower in HEADER_MAP:
            our_field = HEADER_MAP[header_lower]
            if our_field not in mapping:
                mapping[our_field] = col_idx
            else:
                # Already mapped, store as extra
                unmapped.append((header_text, col_idx))
        else:
            # Try partial matching for KN-style long headers
            matched = False
            for pattern, field in HEADER_MAP.items():
                if len(pattern) > 3 and (pattern in header_lower or header_lower.endswith(pattern)):
                    if field not in mapping:
                        mapping[field] = col_idx
                        matched = True
                        break
            if not matched:
                unmapped.append((header_text, col_idx))
    
    return mapping, unmapped


def _detect_direction(ws, column_mapping: dict, header_row: int) -> str:
    if 'direction' in column_mapping:
        for row_idx in range(header_row + 1, min(header_row + 10, ws.max_row + 1)):
            val = ws.cell(row=row_idx, column=column_mapping['direction']).value
            if val:
                v = str(val).strip().lower()
                if 'export' in v: return 'export'
                if 'import' in v: return 'import'
    
    norway_codes = {'osl', 'bgo', 'svg', 'trd', 'bod', 'trf', 'hau', 'tos', 'no'}
    if 'origin' in column_mapping:
        for row_idx in range(header_row + 1, min(header_row + 10, ws.max_row + 1)):
            val = ws.cell(row=row_idx, column=column_mapping['origin']).value
            if val and str(val).strip().lower() in norway_codes:
                return 'export'
    if 'destination' in column_mapping:
        for row_idx in range(header_row + 1, min(header_row + 10, ws.max_row + 1)):
            val = ws.cell(row=row_idx, column=column_mapping['destination']).value
            if val and str(val).strip().lower() in norway_codes:
                return 'import'
    return 'unknown'


def _col_index_to_letter(idx: int) -> str:
    result = ""
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result
